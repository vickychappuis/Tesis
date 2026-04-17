#!/usr/bin/env python3
"""Exporta las tablas de resultados-*.md a un único .xlsx (una hoja por archivo)."""

from __future__ import annotations

import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit("Instalá openpyxl: pip install openpyxl") from e

FILLS = {
    "verde": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "amarillo": PatternFill(fill_type="solid", fgColor="FFEB9C"),
    "rojo": PatternFill(fill_type="solid", fgColor="FFC7CE"),
}

ROOT = Path(__file__).resolve().parent

FILES: list[tuple[str, str]] = [
    ("resultados-acm.md", "ACM"),
    ("resultados-ieee-xplore.md", "IEEE Xplore"),
    ("resultados-sciencedirect.md", "ScienceDirect"),
    ("resultados-springer.md", "Springer"),
    ("resultados-wiley.md", "Wiley"),
]

SEP_RE = re.compile(r"^\|[\s\-:|]+\|\s*$")


def split_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    return [c.strip() for c in line.split("|")]


def extract_tables(md: str) -> list[list[list[str]]]:
    lines = md.splitlines()
    blocks: list[list[str]] = []
    current: list[str] = []
    for line in lines:
        s = line.strip()
        if s.startswith("|") and s.endswith("|"):
            current.append(s)
        else:
            if len(current) >= 2:
                blocks.append(current)
            current = []
    if len(current) >= 2:
        blocks.append(current)

    tables: list[list[list[str]]] = []
    for block in blocks:
        rows: list[list[str]] = []
        for line in block:
            if SEP_RE.match(line):
                continue
            rows.append(split_row(line))
        if rows:
            tables.append(rows)
    return tables


def pick_main_table(tables: list[list[list[str]]]) -> list[list[str]]:
    if not tables:
        return []
    return max(tables, key=lambda t: len(t))


def apply_relevancia_colors(ws) -> None:
    header_row = 1
    relev_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and str(v).strip().lower() == "relevancia":
            relev_col = c
            break
    if relev_col is None:
        return
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=relev_col)
        raw = cell.value
        if raw is None:
            continue
        key = str(raw).strip().lower()
        if key in FILLS:
            cell.fill = FILLS[key]
            cell.value = None


def autosize_columns(ws) -> None:
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 500), values_only=True), start=1):
        max_len = 0
        for cell in col:
            if cell is None:
                continue
            max_len = max(max_len, len(str(cell)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for filename, sheet_title in FILES:
        path = ROOT / filename
        if not path.exists():
            raise FileNotFoundError(path)
        text = path.read_text(encoding="utf-8")
        tables = extract_tables(text)
        rows = pick_main_table(tables)
        if not rows:
            raise ValueError(f"Sin tabla en {filename}")

        ws = wb.create_sheet(title=sheet_title[:31])
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
        apply_relevancia_colors(ws)
        autosize_columns(ws)

    out = ROOT / "resultados.xlsx"
    wb.save(out)
    print(f"Escrito: {out}")


if __name__ == "__main__":
    main()
